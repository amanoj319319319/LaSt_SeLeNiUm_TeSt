import openpyxl

def getRowcount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return (sheet.max_row)

def getColumncount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return (sheet.max_coloumn)

def readData(file,sheetname,rownum,columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return sheet.cell(row=rownum, column=columnnum).value

def writeData(file,sheetname,rownum,columnnum,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    sheet.cell(row=rownum, column=columnnum).value = data
    workbook.save(file)




'''
import xlrd

def getRowcount(file,sheetname):
    #loc = ("C:\\Users\\Manoj\\PycharmProjects\\LastSeleniumTest\\Files\\BOOK2.xls")
    wb = xlrd.open_workbook(file)
    sheet = wb.sheet_by_name(sheetname)
    # Extracting number of rows
    return (sheet.nrows)

def getColumncount(file,sheetname):
    #loc = ("C:\\Users\\Manoj\\PycharmProjects\\LastSeleniumTest\\Files\\BOOK2.xls")
    wb = xlrd.open_workbook(file)
    sheet = wb.sheet_by_name(sheetname)
    return (sheet.ncols)

def readData(file,sheetname,rownum,columnnum):
    #loc = ("C:\\Users\\Manoj\\PycharmProjects\\LastSeleniumTest\\Files\\BOOK2.xls")
    workbook = xlrd.open_workbook(file)
    #workbook = openpyxl.load_workbook(file)
    sheet = workbook.sheet_by_name(sheetname)
    return sheet.cell(rownum,columnnum).value

def writeData(file,sheetname,rownum,columnnum,data):
    #loc = ("C:\\Users\\Manoj\\PycharmProjects\\LastSeleniumTest\\Files\\BOOK2.xls")
    workbook = xlrd.open_workbook(file)
    #workbook = openpyxl.load_workbook(file)
    sheet = workbook.sheet_by_name(sheetname)
    sheet.cell(rownum, columnnum).value = data
    workbook.save(file)

'''







